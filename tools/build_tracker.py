from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
NAVY = "1F3A5F"
MIDBLUE = "2E6CB0"
LIGHT = "DCE6F4"
LIGHT2 = "EFF4FB"
GREEN = "CDE7C4"
AMBER = "FBE3CF"
WHITE = "FFFFFF"
INPUT_BLUE = "0000FF"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def sh(cell, fill=NAVY, color=WHITE, size=11):
    cell.font = Font(name=FONT, bold=True, color=color, size=size)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

def sc(cell, bold=False, fill=None, color="000000", wrap=True, align="left", size=10):
    cell.font = Font(name=FONT, bold=bold, color=color, size=size)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = border

def title_row(ws, text, span, size=14):
    ws.merge_cells(f"A1:{span}1")
    c = ws["A1"]
    c.value = text
    c.font = Font(name=FONT, bold=True, color=WHITE, size=size)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

wb = Workbook()

# ---------------- OVERVIEW ----------------
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False
for i, w in enumerate([28, 62, 18], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
title_row(ws, "Coucou  —  French Progress Tracker", "C", 16)
ws.merge_cells("A2:C2")
c = ws["A2"]
c.value = "Companion to ektasengar.github.io/coucou  ·  pick your target, log your hours, tick milestones"
c.font = Font(name=FONT, italic=True, color="333333", size=10)
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center")

row = 4
ws.merge_cells(f"A{row}:C{row}"); sh(ws[f"A{row}"], fill=MIDBLUE); ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
row += 1
howto = [
    ("1. Set your target", "Decide your goal level on the CEFR scale (A1→C2). For French citizenship you need B2 (since Jan 2026) + the civic test.", ""),
    ("2. Follow the Roadmap", "An editable ~18-month month-by-month path from beginner to B2. Adjust the blue 'Target hrs' to your pace.", ""),
    ("3. Run the Weekly Routine", "The concrete 1–2 h/day plan (40% active / 30% input / 30% speaking).", ""),
    ("4. Use the AI Toolkit tab", "Copy-paste prompts to turn ChatGPT/Claude into a 24/7 tutor; compare the best AI speaking apps.", ""),
    ("5. Log hours weekly", "The Hours Log auto-totals and shows your % toward ~600 h (a realistic B2 estimate).", ""),
    ("6. Tick the Certification list", "Track DELF B2 + (for citizenship) the civic test and paperwork.", ""),
]
for k, v, _ in howto:
    sc(ws[f"A{row}"], bold=True, fill=LIGHT2); ws[f"A{row}"] = k
    sc(ws[f"B{row}"]); ws[f"B{row}"] = v
    sc(ws[f"C{row}"], align="center")
    ws.row_dimensions[row].height = 30
    row += 1

row += 1
ws.merge_cells(f"A{row}:C{row}"); sh(ws[f"A{row}"], fill=MIDBLUE); ws[f"A{row}"] = "CORE PRINCIPLES"
row += 1
for s in [
    "AI is an amplifier, not the engine — pair it with real input, human accountability and immersion.",
    "Output (speaking) is the bottleneck AI finally solves: unlimited, judgement-free practice, 24/7.",
    "'Try first in French, then check' — kill the instant-translate reflex; that's where learning happens.",
    "Consistency beats intensity. 1 focused hour daily > a 7-hour weekend cram.",
    "Aim for the DELF B2 DIPLOMA (permanent) over TCF/TEF tests (valid only 2 years).",
]:
    ws.merge_cells(f"A{row}:C{row}"); sc(ws[f"A{row}"], fill=LIGHT2); ws[f"A{row}"] = "•  " + s
    ws.row_dimensions[row].height = 26
    row += 1

# ---------------- ROADMAP ----------------
ws = wb.create_sheet("Roadmap")
ws.sheet_view.showGridLines = False
headers = ["Month", "Phase", "CEFR", "Main focus", "Action / course", "Milestone", "Target hrs", "Cumulative"]
for i, w in enumerate([13, 15, 9, 33, 28, 24, 11, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
title_row(ws, "18-MONTH ROADMAP  (beginner → B2)", "H")
for i, h in enumerate(headers, 1):
    sh(ws.cell(row=2, column=i, value=h))
ws.row_dimensions[2].height = 28
roadmap = [
    ("Month 1", "Setup", "A0→A1", "Free placement test; install Anki + an AI tutor; phone in French; agree 'French at dinner'.", "Self-study + apps", "Daily habit locked in", 35),
    ("Month 2", "Setup", "A1", "Greetings, present tense, numbers, daily-life vocab. Start AI conversation practice.", "Self-study / app", "Hold a 5-min basic chat", 35),
    ("Month 3", "Setup", "A1", "Consolidate A1; enrol in a structured course (Cours Municipaux / Lingoda / Babbel).", "Enrol in a course", "Course enrolled", 35),
    ("Month 4", "A1→A2", "A1→A2", "Course begins; past tense (passe compose); questions; weekly tutor/tandem.", "Course + tutor", "Course started", 45),
    ("Month 5", "A1→A2", "A2", "Everyday transactions; daily AI role-play (cafe, shop).", "Course + AI", "Order/shop in French", 45),
    ("Month 6", "A1→A2", "A2", "Build vocab breadth; comprehensible-input podcasts (Coffee Break, InnerFrench).", "Course + input", "Follow slow podcasts", 45),
    ("Month 7", "A1→A2", "A2", "Review; optional mock A2; immersion (French films w/ FR subtitles).", "Mock A2", "Solid A2", 45),
    ("Month 8", "A2→B1", "A2→B1", "Future tense, opinions, connectors; AI essay feedback weekly.", "Course term 2", "Express opinions", 45),
    ("Month 9", "A2→B1", "B1", "Narrate & describe; more reading (graded readers, news).", "Course + reading", "Discuss familiar topics", 45),
    ("Month 10", "A2→B1", "B1", "Conditional, subjunctive intro; weekly written paragraphs (AI-corrected).", "Course + writing", "Write short texts", 45),
    ("Month 11", "A2→B1", "B1", "Consolidate B1; start civic-test QCM practice (if citizenship).", "Mock B1", "Solid B1", 45),
    ("Month 12", "B1→B2", "B1→B2", "Shift to OUTPUT: argumentation, debate, structured essays; AI debate partner.", "B2 prep + tutor", "Argue a position", 50),
    ("Month 13", "B1→B2", "B2", "Nuance, complex grammar; learn DELF B2 format (4 papers).", "DELF B2 prep", "Know exam format", 50),
    ("Month 14", "B1→B2", "B2", "Native-speed listening; opinion essays; AI oral-exam simulation.", "DELF B2 prep", "Native-speed listening", 50),
    ("Month 15", "B1→B2", "B2", "Mock DELF B2 #1; (citizenship: take civic test, assemble dossier).", "Mock exam #1", "Mock #1 done", 50),
    ("Month 16", "Certify", "B2", "Target weaknesses; timed essays + oral expose drills.", "DELF B2 prep", "Mock #1 reviewed", 55),
    ("Month 17", "Certify", "B2", "Mock DELF B2 #2; register for the real exam session.", "Register DELF B2", "Mock #2 ≥ pass", 55),
    ("Month 18", "Certify", "B2", "SIT DELF B2. (Citizenship: finalise & submit application.)", "DELF B2 EXAM", "B2 diploma earned", 55),
]
r = 3
first = r
for m, ph, lvl, focus, course, mile, hrs in roadmap:
    sc(ws.cell(row=r, column=1, value=m), bold=True, fill=LIGHT2, align="center")
    sc(ws.cell(row=r, column=2, value=ph), fill=LIGHT2)
    sc(ws.cell(row=r, column=3, value=lvl), align="center", bold=True)
    sc(ws.cell(row=r, column=4, value=focus))
    sc(ws.cell(row=r, column=5, value=course))
    sc(ws.cell(row=r, column=6, value=mile), fill=GREEN)
    sc(ws.cell(row=r, column=7, value=hrs), align="center", color=INPUT_BLUE)
    cc = ws.cell(row=r, column=8)
    cc.value = f"=G{r}" if r == first else f"=H{r-1}+G{r}"
    sc(cc, align="center", bold=True)
    ws.row_dimensions[r].height = 38
    r += 1
sc(ws.cell(row=r, column=1, value="TOTAL"), bold=True, fill=NAVY, color=WHITE, align="center")
for col in range(2, 7):
    sc(ws.cell(row=r, column=col, value=""), fill=NAVY)
sc(ws.cell(row=r, column=7, value=f"=SUM(G{first}:G{r-1})"), bold=True, fill=NAVY, color=WHITE, align="center")
sc(ws.cell(row=r, column=8, value=f"=H{r-1}"), bold=True, fill=NAVY, color=WHITE, align="center")
ws.freeze_panes = "A3"
r += 2
ws.merge_cells(f"A{r}:H{r}")
sc(ws[f"A{r}"], fill=AMBER)
ws[f"A{r}"] = ("'Target hrs' are editable (blue). ~600 guided hrs ≈ a genuine B2; the buffer above 600 covers that self-study "
              "hours are less efficient than guided ones — and real-life immersion is ON TOP of these.")
ws.row_dimensions[r].height = 30

# ---------------- WEEKLY ROUTINE ----------------
ws = wb.create_sheet("Weekly Routine")
ws.sheet_view.showGridLines = False
for col, w in zip("ABC", [20, 50, 18]):
    ws.column_dimensions[col].width = w
title_row(ws, "WEEKLY ROUTINE  (the 1–2 h/day, made concrete)", "C")
r = 3
ws.merge_cells(f"A{r}:C{r}"); sh(ws[f"A{r}"], fill=MIDBLUE); ws[f"A{r}"] = "DAILY SPLIT  (≈ 40 / 30 / 30)"
r += 1
for i, h in enumerate(["Block", "What to do", "Share"], 1):
    sh(ws.cell(row=r, column=i, value=h))
r += 1
for a, b, c in [
    ("Active study", "Course homework + grammar + Anki flashcards (spaced repetition).", "~40%"),
    ("Input", "Comprehensible input slightly above your level: TV5Monde, RFI slow-news podcast, Coffee Break / InnerFrench.", "~30%"),
    ("Output (speak)", "AI voice tutor daily + 1 weekly human session (italki/tandem). The part most people skip.", "~30%"),
]:
    sc(ws.cell(row=r, column=1, value=a), bold=True, fill=LIGHT2)
    sc(ws.cell(row=r, column=2, value=b))
    sc(ws.cell(row=r, column=3, value=c), align="center", bold=True, fill=GREEN)
    ws.row_dimensions[r].height = 38
    r += 1
r += 1
ws.merge_cells(f"A{r}:C{r}"); sh(ws[f"A{r}"], fill=MIDBLUE); ws[f"A{r}"] = "WEEKLY RHYTHM"
r += 1
for i, h in enumerate(["Day", "Plan", "~Time"], 1):
    sh(ws.cell(row=r, column=i, value=h))
r += 1
for d, p, t in [
    ("Monday", "Structured course / lesson", "1–2 h"),
    ("Tuesday", "Self-study: Anki + grammar + input", "1 h"),
    ("Wednesday", "AI conversation + course", "1–2 h"),
    ("Thursday", "Speaking session — human tutor or tandem", "1 h"),
    ("Friday", "Light: podcast on commute + Anki", "0.5–1 h"),
    ("Saturday", "Immersion: French film w/ FR subtitles, or news", "1–1.5 h"),
    ("Sunday", "Review week + write a short AI-corrected paragraph", "1 h"),
    ("Every day", "'Try first, then translate' in real life + 'French zone' at home", "passive"),
]:
    fill = AMBER if d == "Every day" else LIGHT2
    sc(ws.cell(row=r, column=1, value=d), bold=True, fill=fill)
    sc(ws.cell(row=r, column=2, value=p), fill=(AMBER if d == "Every day" else None))
    sc(ws.cell(row=r, column=3, value=t), align="center", fill=(AMBER if d == "Every day" else None))
    ws.row_dimensions[r].height = 24
    r += 1

# ---------------- AI TOOLKIT & PROMPTS ----------------
ws = wb.create_sheet("AI Toolkit & Prompts")
ws.sheet_view.showGridLines = False
for col, w in zip("ABC", [24, 70, 18]):
    ws.column_dimensions[col].width = w
title_row(ws, "AI TOOLKIT & PROMPT COOKBOOK", "C")
r = 3
ws.merge_cells(f"A{r}:C{r}"); sh(ws[f"A{r}"], fill=MIDBLUE); ws[f"A{r}"] = "BEST AI SPEAKING APPS (2026)"
r += 1
for i, h in enumerate(["Tool", "Best for / notes", "Price"], 1):
    sh(ws.cell(row=r, column=i, value=h))
r += 1
for a, b, c in [
    ("Langua", "Best overall for French: natural voices, deep conversation, detailed correction reports.", "Freemium / paid"),
    ("Speak", "Best for beginners & structured daily practice (note: lenient pronunciation scoring).", "~$20/mo"),
    ("ChatGPT (Advanced Voice)", "Flexible all-rounder; great for free-form role-play and feedback.", "~$20/mo"),
    ("Praktika", "Affordable AI avatars, immersive practice; lighter feedback.", "~$8/mo"),
    ("Claude (text)", "Excellent text tutor & prompt engine (voice mode is English-only in early 2026).", "Free / paid"),
    ("Watch-outs", "AI can be over-lenient (false confidence) and occasionally output wrong French — verify with input + a human.", "—"),
]:
    sc(ws.cell(row=r, column=1, value=a), bold=True, fill=LIGHT2)
    sc(ws.cell(row=r, column=2, value=b))
    sc(ws.cell(row=r, column=3, value=c), align="center")
    ws.row_dimensions[r].height = 32
    r += 1
r += 1
ws.merge_cells(f"A{r}:C{r}"); sh(ws[f"A{r}"], fill=MIDBLUE); ws[f"A{r}"] = "COPY-PASTE PROMPTS  (more on the website)"
r += 1
for i, h in enumerate(["Prompt name", "Paste this into ChatGPT / Claude", "Use it for"], 1):
    sh(ws.cell(row=r, column=i, value=h))
r += 1
prompts = [
    ("24/7 tutor (system)", "You are my patient French tutor. I'm level [A2]. Speak mostly in simple French. After each of my messages: (1) reply naturally, (2) gently correct my mistakes with a short why, (3) suggest one better phrase. Keep me talking with questions. Don't let me switch to English.", "Daily conversation"),
    ("Cafe role-play", "Role-play: you are a waiter in a Paris cafe, I'm ordering. Stay in French at my level [A2]. If I'm stuck, offer the phrase. At the end, list my 3 most useful corrections.", "Real-life practice"),
    ("Prefecture interview", "Simulate a French naturalisation/prefecture interview in French. Ask typical questions about my life, integration and French values. Then give feedback on fluency and what to improve.", "Citizenship prep"),
    ("Correct my essay", "Here is my French text: [paste]. Correct it and show a table: my version | correction | why. Then rewrite it cleanly at B2 level.", "DELF B2 writing"),
    ("Graded reader", "Write a 150-word French story at level [B1] about [topic]. Use ~90% common words, 10% new. After it, list the new words with English meanings.", "Reading + vocab"),
    ("Anki cards (CSV)", "From this text [paste], extract 15 useful words/phrases as CSV: french,english,example_sentence. Output only the CSV.", "Vocabulary"),
    ("Grammar explainer", "Explain [the subjunctive] simply with 5 example sentences and a 3-question quiz. Wait for my answers before marking.", "Targeted grammar"),
    ("EN->FR daily drill", "Give me 5 English sentences at [A2] to translate to French. After I answer, mark each, correct errors, explain briefly.", "Active recall"),
    ("DELF B2 oral sim", "Run a DELF B2 oral exam: give me a short article, I argue a position for 3 min, then you ask 2 follow-ups and debate. Score me on the DELF grid and give tips.", "Exam rehearsal"),
    ("Level rewrite", "Rewrite this sentence three ways — A2, B1, B2 — and label each: [sentence].", "See progress"),
]
for a, b, c in prompts:
    sc(ws.cell(row=r, column=1, value=a), bold=True, fill=LIGHT2)
    sc(ws.cell(row=r, column=2, value=b))
    sc(ws.cell(row=r, column=3, value=c), align="center")
    ws.row_dimensions[r].height = 56
    r += 1

# ---------------- RESOURCES & COST ----------------
ws = wb.create_sheet("Resources & Cost")
ws.sheet_view.showGridLines = False
for col, w in zip("ABCD", [26, 38, 18, 30]):
    ws.column_dimensions[col].width = w
title_row(ws, "RESOURCES & COST  (cheapest → premium)", "D")
for i, h in enumerate(["Resource", "What it's for", "Cost", "Tip"], 1):
    sh(ws.cell(row=2, column=i, value=h))
ws.row_dimensions[2].height = 22
res = [
    ("Immersion / talking to people", "Daily real-world practice — your biggest free asset", "Free", "'Try first, then translate'"),
    ("TV5Monde 'Apprendre le francais'", "Graded listening/reading by CEFR level", "Free", "apprendre.tv5monde.com"),
    ("RFI 'Journal en francais facile'", "Daily slow-news podcast", "Free", "Great for commutes"),
    ("InnerFrench / Coffee Break French", "Comprehensible-input podcasts", "Free", "Best for A2→B1"),
    ("Anki", "Vocabulary via spaced repetition", "Free", "10 min/day; AI builds the cards"),
    ("An AI speaking app (Langua/Speak)", "Unlimited speaking practice + feedback", "Free–$20/mo", "The modern superpower"),
    ("italki", "1-on-1 lessons with native tutors", "~$8–$30/hr", "1–2x per week"),
    ("Busuu / Babbel", "Structured grammar + writing feedback", "~$10–14/mo", "Babbel plateaus ~B1"),
    ("Lingoda", "Live CEFR-aligned group classes", "Pricier", "Good for structure"),
    ("Cours Municipaux d'Adultes (Paris)", "Cheap in-person structured course", "~€50–350/term", "If you're in Paris"),
    ("Alliance Francaise", "Quality in-person courses + DELF centre", "Mid–premium", "Worldwide"),
    ("DELF B2 exam", "The PERMANENT diploma", "~€150–280 (once)", "Fund via CPF if employed in FR"),
]
r = 3
for a, b, c, d in res:
    sc(ws.cell(row=r, column=1, value=a), bold=True, fill=LIGHT2)
    sc(ws.cell(row=r, column=2, value=b))
    sc(ws.cell(row=r, column=3, value=c), align="center")
    sc(ws.cell(row=r, column=4, value=d))
    ws.row_dimensions[r].height = 26
    r += 1

# ---------------- CERTIFICATION ----------------
ws = wb.create_sheet("Certification")
ws.sheet_view.showGridLines = False
for col, w in zip("ABCD", [30, 46, 16, 14]):
    ws.column_dimensions[col].width = w
title_row(ws, "CERTIFICATION & CITIZENSHIP CHECKLIST", "D")
for i, h in enumerate(["Step", "Details", "Target date", "Status"], 1):
    sh(ws.cell(row=2, column=i, value=h))
ws.row_dimensions[2].height = 22
steps = [
    ("Choose exam: DELF B2", "Permanent diploma (never expires). Preferred over TCF/TEF IRN (valid 2 yrs).", "", "To do"),
    ("Reach genuine B2", "Both oral AND written — see Roadmap.", "", "In progress"),
    ("Book + sit DELF B2", "4 papers: listening, reading, writing, speaking.", "", "Not started"),
    ("— For citizenship below —", "", "", ""),
    ("5 yrs continuous residence", "Avoid long absences that break continuity.", "", ""),
    ("Civic test", "40 QCM on history/values/culture; pass = 32/40 (required since 2026).", "", "Not started"),
    ("Gather documents", "Birth cert (apostille + sworn translation), passport, 3 yrs tax notices, proof of address, payslips.", "", "Not started"),
    ("Submit naturalisation dossier", "Online via ANEF / prefecture, once B2 + civic test + docs ready.", "", "Not started"),
    ("Assimilation interview", "Prefecture interview in French on values & integration.", "", "Not started"),
]
r = 3
for a, b, c, d in steps:
    section = a.startswith("—")
    sc(ws.cell(row=r, column=1, value=a), bold=True, fill=(MIDBLUE if section else LIGHT2), color=(WHITE if section else "000000"))
    sc(ws.cell(row=r, column=2, value=b), fill=(MIDBLUE if section else None))
    sc(ws.cell(row=r, column=3, value=c), align="center", color=INPUT_BLUE, fill=(MIDBLUE if section else None))
    sc(ws.cell(row=r, column=4, value=d), align="center", fill=(MIDBLUE if section else AMBER), color=(WHITE if section else "000000"))
    ws.row_dimensions[r].height = 32
    r += 1
r += 1
ws.merge_cells(f"A{r}:D{r}")
sc(ws[f"A{r}"], fill=AMBER, bold=True)
ws[f"A{r}"] = "Verify the latest rules on service-public.fr before filing — requirements have tightened recently and may change again."
ws.row_dimensions[r].height = 28

# ---------------- HOURS LOG ----------------
ws = wb.create_sheet("Hours Log")
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [16, 14, 16, 18, 40]):
    ws.column_dimensions[col].width = w
title_row(ws, "HOURS LOG  (what gets measured gets done)", "E")
for i, h in enumerate(["Week", "Hours", "Cumulative", "% toward B2", "Notes"], 1):
    sh(ws.cell(row=2, column=i, value=h))
ws.row_dimensions[2].height = 24
GOAL_CELL = "H1"
ws[GOAL_CELL] = 600
ws["G1"] = "B2 goal (hrs):"
sc(ws["G1"], bold=True, fill=LIGHT2, align="right")
sc(ws[GOAL_CELL], bold=True, align="center", color=INPUT_BLUE)
first = 3
n = 24
for i in range(n):
    r = first + i
    sc(ws.cell(row=r, column=1, value=f"Week {i+1}"), bold=True, fill=LIGHT2, align="center")
    hcell = ws.cell(row=r, column=2, value=0); sc(hcell, align="center", color=INPUT_BLUE)
    cum = ws.cell(row=r, column=3)
    cum.value = f"=B{r}" if i == 0 else f"=C{r-1}+B{r}"
    sc(cum, align="center", bold=True)
    pct = ws.cell(row=r, column=4, value=f"=C{r}/$H$1")
    sc(pct, align="center"); pct.number_format = "0.0%"
    sc(ws.cell(row=r, column=5, value=""))
    ws.row_dimensions[r].height = 18
r = first + n
sc(ws.cell(row=r, column=1, value="TOTAL"), bold=True, fill=NAVY, color=WHITE, align="center")
sc(ws.cell(row=r, column=2, value=f"=SUM(B{first}:B{r-1})"), bold=True, fill=NAVY, color=WHITE, align="center")
sc(ws.cell(row=r, column=3, value=f"=C{r-1}"), bold=True, fill=NAVY, color=WHITE, align="center")
pct = ws.cell(row=r, column=4, value=f"=C{r}/$H$1"); sc(pct, bold=True, fill=NAVY, color=WHITE, align="center"); pct.number_format = "0.0%"
sc(ws.cell(row=r, column=5, value=""), fill=NAVY)
ws.freeze_panes = "A3"

import os
SITE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
out = os.path.join(SITE, "assets/French_Learning_Tracker.xlsx")
wb.save(out)
print("saved", out, os.path.getsize(out), "bytes")
